from pathlib import Path

import pandas as pd

from dependency_metrics.reporting import (
    export_osv_data,
    export_per_release_worksheets,
    export_worksheets,
    safe_filename_stem,
    safe_sheet_name,
    save_results_json,
)
from openpyxl import load_workbook


def test_reporting_exports(tmp_path: Path):
    output_dir = tmp_path / "out"
    dep_df = pd.DataFrame({"dependency": ["demo"], "value": [1]})
    results = {
        "ttu": 1.0,
        "ttr": 2.0,
        "num_dependencies": 1,
        "dependency_data": {"demo": dep_df},
        "osv_data": pd.DataFrame({"package": ["demo"], "vul_id": ["CVE-1"]}),
    }

    results_file = save_results_json(results, output_dir, "demo")
    osv_file = export_osv_data(results, output_dir, "demo")
    excel_file = export_worksheets(results, output_dir, "demo")

    assert results_file.exists()
    assert osv_file is not None and osv_file.exists()
    assert excel_file is not None and excel_file.exists()


def test_scoped_npm_package_name_does_not_create_subdirectory(tmp_path: Path):
    """Scoped npm packages like "@medv/finder" contain a "/" — naively
    interpolating them into a path makes pandas/Path treat it as a (missing)
    subdirectory ("output/@medv/finder_results.json"), raising
    "Cannot save file into a non-existent directory". Exporters must sanitize
    the package name into a single filename component, written directly inside
    output_dir (no nested directory created)."""
    output_dir = tmp_path / "out"
    package = "@medv/finder"
    results = {"ttu": 1.0, "ttr": 2.0, "num_dependencies": 0}

    results_file = save_results_json(results, output_dir, package)

    assert results_file.exists()
    assert results_file.parent == output_dir
    assert not (output_dir / "@medv").exists()
    assert "/" not in results_file.name


def test_safe_filename_stem_replaces_path_separators():
    assert safe_filename_stem("@medv/finder") == "@medv__finder"
    assert "/" not in safe_filename_stem("@scope/pkg")
    assert "\\" not in safe_filename_stem("scope\\pkg")
    assert safe_filename_stem("plain-package") == "plain-package"


def test_export_worksheets_sanitizes_scoped_dependency_name(tmp_path: Path):
    """A dependency named "@scope/pkg" contains "/" — an Excel-forbidden sheet
    title character (openpyxl raises "Invalid character / found in sheet
    title" otherwise). The exported sheet must use a sanitized title."""
    output_dir = tmp_path / "out"
    output_dir.mkdir(parents=True)
    dep_df = pd.DataFrame({"dependency": ["@scope/pkg"], "value": [1]})
    results = {
        "ttu": 1.0,
        "ttr": 2.0,
        "num_dependencies": 1,
        "dependency_data": {"@scope/pkg": dep_df},
    }

    excel_file = export_worksheets(results, output_dir, "demo")

    assert excel_file is not None and excel_file.exists()
    wb = load_workbook(excel_file)
    assert "@scope/pkg" not in wb.sheetnames
    assert "@scope_pkg" in wb.sheetnames


def test_export_per_release_worksheets_sanitizes_scoped_dependency_name(tmp_path: Path):
    output_dir = tmp_path / "out"
    output_dir.mkdir(parents=True)
    frame_df = pd.DataFrame(
        {
            "dependency": ["@scope/pkg"],
            "interval_start": [pd.Timestamp("2024-01-01", tz="UTC")],
            "interval_end": [pd.Timestamp("2024-02-01", tz="UTC")],
        }
    )
    release_results = [
        {
            "summary": {"package_version": "1.0.0", "mttu": 1.0, "mttr": 0.0},
            "dependency_frames": [frame_df],
        }
    ]

    excel_file = export_per_release_worksheets(release_results, output_dir, "demo")

    assert excel_file is not None and excel_file.exists()
    wb = load_workbook(excel_file)
    assert "@scope/pkg" not in wb.sheetnames
    assert any("@scope_pkg" in name for name in wb.sheetnames)


def test_safe_sheet_name_replaces_invalid_characters_and_truncates():
    assert safe_sheet_name("@scope/pkg") == "@scope_pkg"
    for ch in "[]:*?/\\":
        assert ch not in safe_sheet_name(f"weird{ch}name")
    long_name = "a" * 40
    assert len(safe_sheet_name(long_name)) == 31
    assert len(safe_sheet_name(long_name, "_PR")) == 31
    assert safe_sheet_name(long_name, "_PR").endswith("_PR")
